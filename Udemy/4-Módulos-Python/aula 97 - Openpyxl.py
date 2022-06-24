"""Openpyxl - Planilhas do Excel em Python

Openpyxl é uma biblioteca Python para leitura e gravação de arquivos Excel
(com extensão xlsx / xlsm / xltx / xltm).
O módulo openpyxl permite que o programa Python leia e modifique arquivos Excel.


-https://openpyxl.readthedocs.io/en/stable/
-https://acervolima.com/lendo-um-arquivo-excel-usando-o-modulo-python-openpyxl/
"""
import openpyxl
from random import uniform

# carrega o excel
requests = openpyxl.load_workbook(
    r'C:\Users\Igor\Desktop\Estudos\Programação-em-Python\Mundo-invertido\Files\excel\pedidos.xlsx')

# Mostra quantas guias tem no arquivo
worksheet_name = requests.sheetnames  # Nome da planilha
spreadsheet1 = requests['Page 1']  # planilha1

print(spreadsheet1['b4'].value)  # Exibi um valor selecionado do excel

for field in spreadsheet1['b']:  # Itera uma coluna
    print(field.value)

for line in spreadsheet1['a1:c2']:
    for column in line:
        print(column.value)

for line in spreadsheet1:
    if line[0].value is not None:
        print(line[0].value, end=' ')

    if line[1].value is not None:
        print(line[1].value, end=' ')

    if line[2].value is not None:
        print(line[2].value)

spreadsheet1['b3'].value = 2200


for line in range(5, 16):
    order_number = line - 1
    spreadsheet1.cell(line, 1).value = order_number
    spreadsheet1.cell(line, 2).value = 1200 + line

    price = round(uniform(10, 100), 2)
    spreadsheet1.cell(line, 3).value = price


requests.save('new_worksheet.xlsx')  # Salva o arquivo


new_spreadsheet0 = openpyxl.Workbook()
new_spreadsheet0.create_sheet('Page 1', 0)
new_spreadsheet0.create_sheet('Page 2', 0)

new_spreadsheet1 = new_spreadsheet0['Page 1']
new_spreadsheet2 = new_spreadsheet0['Page 2']

# planilha 1
for line in range(1, 11):
    order_number = line - 1
    spreadsheet1.cell(line, 1).value = order_number
    spreadsheet1.cell(line, 2).value = 1200 + line

    price = round(uniform(10, 100), 2)
    spreadsheet1.cell(line, 3).value = price

# planilha 2
for line in range(1, 11):
    new_spreadsheet2.cell(
        line, 1).value = f'Igor {line} {round(uniform(10,100),2)}'
    new_spreadsheet2.cell(
        line, 2).value = f'Michele {line} {round(uniform(10,100),2)}'
    new_spreadsheet2.cell(
        line, 3).value = f'Alice {line} {round(uniform(10,100),2)}'

new_spreadsheet0.save('new_spreadsheet.xlsx')
